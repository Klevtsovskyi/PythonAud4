"""
Скласти програму, яка працює в оточенні веб-сервера, для конвертування
валют. Поточні курси валют містяться у файлі Excel на сервері у форматі:
<код валюти 1> <код валюти 2> <курс>. Код валюти – це рядок з 3 символів,
наприклад UAH, USD, EUR тощо.
Currency1 Currency2 Rate
USD       UAH       24,81
EUR       UAH       27,75
…	      …	        …
Програма повинна забезпечити вибір двох валют зі списків кодів валют на
сторінці, введення суми у першій валюті та показ суми у другій валюті.

b) використати WSGI
"""

import cgi
import openpyxl
from string import Template

OPTION = """
      <option value="$cur">$cur</option>
"""


class ExchangeRate:

    def __init__(self, data):
        self._data = data

    def get_currencies(self):
        wb = openpyxl.load_workbook(self._data)
        ws = wb.worksheets[0]
        return [ws.cell(i, 1).value for i in range(1, ws.max_row+1)]

    def get_exchange_rates(self):
        wb = openpyxl.load_workbook(self._data)
        ws = wb.worksheets[1]
        return [[cell.value for cell in row] for row in ws.rows][1:]

    def obtain_rate(self, cur1, cur2, amount):
        if cur1 == cur2:
            return amount
        for c1, c2, rate in self.get_exchange_rates():
            if c1 == cur1 and c2 == cur2:
                return amount * rate
            elif c1 == cur2 and c2 == cur1:
                return amount / rate

    def __call__(self, environ, start_response):
        path = environ.get("PATH_INFO", "").lstrip("/")
        params = {"currencies": "", "result": ""}
        status = "200 OK"
        headers = [("Content-Type", "text/html; charset=utf-8")]

        # http://127.0.0.1:8000/
        if path == "":
            currencies = ""
            for cur in self.get_currencies():
                currencies += Template(OPTION).substitute(cur=cur)
            params["currencies"] = currencies
            html = "templates/currencies.html"

        # http://127.0.0.1:8000/exchange_rate
        elif path == "exchange_rate":
            form = cgi.FieldStorage(fp=environ["wsgi.input"], environ=environ)
            cur1 = form.getfirst("from", "")
            cur2 = form.getfirst("to", "")
            amount = form.getfirst("amount", "")
            if cur1 and cur2 and amount:
                rate = round(self.obtain_rate(cur1, cur2, float(amount)), 2)
                params["result"] = "{} {} = {} {}".format(amount, cur1, rate, cur2)
            # Якщо у формі задано недостатньо параметрів, перенаправляємо на головну сторінку
            else:
                status = "303 SEE OTHER"
                headers.append(("Location", "/"))
            html = "templates/exchange_rate.html"

        # http://127.0.0.1:8000/<будь-який інший запит>
        else:
            status = "404 NOT FOUND"
            html = "templates/error_404.html"

        start_response(status, headers)
        with open(html, encoding="utf-8") as f:
            page = Template(f.read()).substitute(params)
        return [bytes(page, encoding="utf-8")]


HOST = ""
PORT = 8000

if __name__ == '__main__':
    app = ExchangeRate("data/currencies.xlsx")
    from wsgiref.simple_server import make_server
    print(" === Local webserver === ")
    make_server(HOST, PORT, app).serve_forever()
